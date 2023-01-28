# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import numpy as np
from openpyxl import Workbook, load_workbook
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

wb = load_workbook('Beetroot_Final_project.xlsx',data_only=True)
# Завантажуємо дані у сторінку вихідних даних
ws = wb['Свод_таблица']
row_count = ws.max_row

# Вихідні дані для підстановки у комірки:
a_range = np.linspace(0.02, 0.1, num=10)
b_range = np.linspace(0.1, 0.3, num=10)
c_range = np.linspace(0.1, 0.2, num=5)
d_range = np.linspace(0.1, 0.4, num=10)
e_range = np.linspace(0.02, 0.1, num=10)


data = [ws.cell(row=26, column=n).value for n in range(3, 25)]
for row in range(27, row_count+1):
    for i in a_range:
        ws.cell(row=26, column=3).value = i
    for j in b_range:
        ws.cell(row=26, column=4).value = j
    for k in c_range:
        ws.cell(row=26, column=5).value = k
    for l in d_range:
        ws.cell(row=26, column=6).value = l
    for m in e_range:
        ws.cell(row=26, column=7).value = m
    for col in range(3, 25):
        ws.cell(row=row, column=col).value = ws.cell(row=26, column=col).value
wb.save('Beetroot_Final_project.test.xlsx')
# # Підстановка даних у комірки С26:G26
# # for i in range (5):
# for i,j in enumerate(np.linspace(0.2, 0.5, num=5)):
#     print(i,j)
#     ws.cell(row=26, column=3 + i).value = j
#     data = [ws.cell(row=26, column=i).value for i in range(3, 25)]
#         # Copying data
# for cell in data:
#     ws.cell(row=26+i, column=i).value for i in range(3, 25) = ws.cell(row=26, column=3 + i).value = j
# # Копіювання даних з рядка С26:Х26
# for k in range(27, row_count + 1):
#     # Отримуємо дані з 26-го рядка
#     data = ws.cell(row=26, column=3).value
#     # Вставляємо дані в наступний рядок
#     ws.cell(row=k, column=1).value = data

# print(np.linspace(0.2, 0.5, num=5))
# print(round(ws['J26'].value,3))
# print(data)
# print(range(27, row_count + 1))

# Save the workbook

# # value = np.linspace(0.2,0.5,num = 5)
# # value = range(0.2, 0.5, 0.05)
# def make_calculation (value):
#     last_row = wb['Свод_таблица'].max_row
#     # cells = ws.getcells(['C26'],['D26'],['E26'],['F26'],['G26'])
#     # result_page = ws
#         # value = range(0.2, 0.5, 0.05)
#     for i in cells:
#         for j in value:
#             ws.cell(row=last_row + 1, column=3).value = j
#         last_row += 1
# make_calculation (value)
# print(value)


import openpyxl

# Load the workbook
# workbook = openpyxl.load_workbook('example.xlsx')

# Select the worksheet
# worksheet = workbook['Sheet1']
#
# # Get the current number of rows
# num_rows = worksheet.max_row
#
# for i in range(10):
#     # Input the value from the for loop into cells C26:G26
#     for j in range(5):
#         worksheet.cell(row=num_rows+1, column=3+j).value = i
#
#     # Perform calculations based on formulas embedded in cells H26:X26
#     # This will update the values in H26:X26
#     worksheet.calculate_formulas()
#     workbook.save('example.xlsx')
#     # Copy the values in row C26:X26 to the next row C27:G27
#     for col_num in range(3,14):
#         worksheet.cell(row=num_rows+2, column=col_num).value = worksheet.cell(row=num_rows+1, column=col_num).value
#     num_rows +=1
# # save the changes
# workbook.save('example.xlsx')

